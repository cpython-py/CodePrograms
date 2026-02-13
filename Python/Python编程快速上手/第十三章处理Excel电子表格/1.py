import openpyxl as opx


wb = opx.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

